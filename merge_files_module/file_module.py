from pathlib import Path
from datetime import datetime

from loguru import logger
from openpyxl import Workbook, load_workbook


class FileModule:
    def __init__(self):
        self.merge_files_names: dict[str, list[str]] = {
            # --------------------- ДКП ---------------------
            "ДКП RESULT": [
                r"C:\Users\Projects\my_projects\parser_3000\p3000\parsers\Avito_parser\all_exel\exel_2026-17-10\2026-07-16_Ivanovo.xlsx",
                r"C:\Users\Projects\my_projects\parser_3000\p3000\parsers\Avito_parser\all_exel\exel_2026-17-10\2026-07-16_Kovrov.xlsx",
                r"C:\Users\Projects\my_projects\parser_3000\p3000\parsers\Avito_parser\all_exel\exel_2026-17-10\2026-07-16_Vladimir.xlsx",
            ],

            # ------------------ ОМЦ Владимир ------------------
            # "ОМЦ Владимир RESULT": [
            #     r"...",
            #     r"...",
            # ],

            # ------------------ ОМЦ Иваново ------------------
            # "ОМЦ Иваново RESULT": [
            #     r"...",
            #     r"...",
            # ],
        }

        self.file_name: dict[str, str] = {
            "ОМЦ Владимир RESULT": (
                r"C:\Users\Projects\my_projects\parser_3000\merge_files_module\нити_макет.xlsx"
            ),

            # "ОМЦ Иваново RESULT": (
            #     r"C:\Users\Projects\my_projects\parser_3000\merge_files_module\другой_макет.xlsx"
            # ),
        }

        self.add_column_data: dict = {
            "column_name": "Дата",
            "column_data": "19.07.2026",
        }


    def _get_result_folder(self) -> Path:
        today = datetime.now().strftime("%d.%m.%Y")

        result_folder = Path.cwd() / "result_files" / today
        result_folder.mkdir(parents=True, exist_ok=True)

        logger.info(f"Папка результатов: {result_folder}")

        return result_folder


    def add_first_column(self):
        result_folder = self._get_result_folder()

        for result_name, file_path in self.file_name.items():

            logger.info(f"Начата обработка файла '{file_path}'")

            try:
                wb = load_workbook(file_path)
                ws = wb.active

                logger.debug("Вставка первого столбца")

                ws.insert_cols(1)

                ws["A1"] = self.add_column_data["column_name"]

                for row in range(2, ws.max_row + 1):
                    ws.cell(
                        row=row,
                        column=1,
                        value=self.add_column_data["column_data"]
                    )

                save_path = result_folder / f"{result_name}.xlsx"

                wb.save(save_path)

                logger.success(
                    f"Файл '{result_name}.xlsx' успешно сохранён"
                )

            except Exception:
                logger.exception(
                    f"Ошибка при обработке файла '{file_path}'"
                )
                raise


    def merge_excel_files(self):
        logger.info(
            f"Найдено групп для объединения: {len(self.merge_files_names)}"
        )

        result_folder = self._get_result_folder()

        try:
            for result_name, files in self.merge_files_names.items():

                logger.info(
                    f"Начато объединение группы '{result_name}' "
                    f"({len(files)} файлов)"
                )

                new_wb = Workbook()
                new_ws = new_wb.active

                first_file = True
                current_row = 1

                for file_path in files:

                    logger.info(f"Чтение файла: {file_path}")

                    wb = load_workbook(file_path)
                    ws = wb.active

                    start_row = 1 if first_file else 2

                    copied_rows = 0

                    for row in ws.iter_rows(
                        min_row=start_row,
                        max_row=ws.max_row,
                        values_only=True,
                    ):
                        for col, value in enumerate(row, start=1):
                            new_ws.cell(
                                row=current_row,
                                column=col,
                                value=value,
                            )

                        current_row += 1
                        copied_rows += 1

                    logger.debug(
                        f"Из '{Path(file_path).name}' "
                        f"скопировано строк: {copied_rows}"
                    )

                    first_file = False

                save_path = result_folder / f"{result_name}.xlsx"

                new_wb.save(save_path)

                logger.success(
                    f"Группа '{result_name}' успешно объединена "
                    f"в '{save_path.name}'"
                )

        except Exception:
            logger.exception("Ошибка при объединении Excel файлов")
            raise


if __name__ == "__main__":
    logger.add(
        "logs/file_module.log",
        rotation="10 MB",
        retention="30 days",
        compression="zip",
        encoding="utf-8",
        level="DEBUG",
    )

    fm = FileModule()

    # Добавление столбца во все файлы из self.file_name
    # fm.add_first_column()

    # Объединение всех групп из self.merge_files_names
    fm.merge_excel_files()